from openpyxl import Workbook
from docx import Document
from docx.shared import Inches
from docx.enum.text import WD_ALIGN_PARAGRAPH
import textract
import re
from textwrap import wrap
from test.test_pyexpat import PositionTest
from docx.text.paragraph import Paragraph

wb = Workbook()
ws = wb.active

document = Document('bc1.docx')
paragpraphs = document.paragraphs
index = 3
prompt_index = 1
label_count = 0
is_bold = False
ask_string = "ASK"
code3_string = "CODE 3"
ask_string = "ASK"
tip_string = "TIP"
b_string = "B"
c_string = "C"
colon_string = ":"
module_name = "PSYC"
excel_alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 
            'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 
            'AW', 'AX', 'AY', 'AZ']
text = ''
bold_turn = False
ask_turn = False
tip_turn = False
labled = False
lable_one = False
lable_two = False
dropdownText = ["? - insufficient information", "1 - Absent / False", "2 - Subthreshold", "3 - Threshold / True"] 
dropdownNumber = ["0", "1", "2", "3"]
dropdownLabel = ["Dropdown", "", "", ""]
new_strip = ""

def name_generator_codeask(question_text):
    capital_letters = ''
    module_name = 'PSYC'
    specific_module_name = raw_input("Module Name")
    if question_text.re.match(b_string+'\d'+'\W', question_text.lstrip(), flags = 0):
        capital_letters = question_text[0]+question_text[1]   
    return module_name+specific_module_name+"_"+capital_letters

def clear(number):
    number = 0
    
# distribute text into according columns and divide it into multiple lines
ws['A1'] = "Position"
ws['B1'] = "Prompt"
ws['C1'] = "Name"
ws['D1'] = "Type"
ws['E1'] = "Length"
ws['F1'] = "Code"
ws['G1'] = "Description".ljust(50)
ws['K1'] = "IS_HEADER"
ws['L1'] = "Required"
ws['M1'] = "Comment"
ws['B2'] = "Module A: \nMood Disorder Episodes"

def two_or_three(number):
    if number < 102:
        return 2
    else: 
        return 3
 
for paragpraph in paragpraphs:
    #bold
    if re.match(b_string+'\d+'+'\W', paragpraph.text.lstrip() , flags=0) or re.match(c_string+'\d+'+'\W', paragpraph.text.lstrip() , flags=0):
        index += 1
        label_count += 1
        text = "<table> \n<tr> \n<td colspan=2 valign = top width = 700><b>" + (paragpraph.text.lstrip()) + "</b><p></p></td> \n</tr> \n<tr>"
        ws[excel_alphabet[6]+str(index)] ='\n'.join(['\n'.join(wrap(block, width=50)) for block in text.splitlines()])
        ws[excel_alphabet[3]+str(index)] = "Label"
        ws[excel_alphabet[2]+str(index)] = "ModA_" + paragpraph.text.lstrip().partition(":")[0] + "_Label"
        if (lable_one == False):
            ws[excel_alphabet[1]+str(index)] = paragpraph.text.lstrip().partition(":")[0] + "_Label"
            new_strip = paragpraph.text.lstrip().partition(":")[0]
            labled_one = True
        '''
        if labled == False:
            ws[excel_alphabet[1]+str(index)] = paragpraph.text.lstrip().partition(":")[0] + "_Label" 
            ws[excel_alphabet[1]+str(index+1)] = paragpraph.text.lstrip().partition(":")[0] + "_Label" 
            ws[excel_alphabet[1]+str(index+2)] = paragpraph.text.lstrip().partition(":")[0]  + "_Label"
            ws[excel_alphabet[1]+str(index+4)] = paragpraph.text.lstrip().partition(":")[0]  + "_Label_2" 
            labled = True
        '''
    #code3
    if paragpraph.text.lstrip().find(code3_string.lstrip()) == 0:
        index += 1
        label_count += 1
        text = "<td valign = top width = 350 <b> Code: </b><p>" + paragpraph.text.lstrip() + "<p/></td>"
        ws[excel_alphabet[6] + str(index)] = '\n'.join(['\n'.join(wrap(block, width=50)) for block in text.splitlines()])
        #ws[excel_alphabet[3]+str(index)] = "Label"
    #ask
    if paragpraph.text.lstrip().find(ask_string.lstrip()) == 0:
        index += 1
        label_count += 1
        text = "<td  valign=top width=350><b>Ask:</b><p>" + paragpraph.text.lstrip() + "</p></td>" + "\n</tr>\n</table>"
        ws[excel_alphabet[6]+str(index)] = '\n'.join(['\n'.join(wrap(block, width=50)) for block in text.splitlines()])
        #ws[excel_alphabet[3]+str(index)] = "Label"
        ask_turn = True
    #drop down 
    i = 0
    if ask_turn == True:
        for i in range(4):
            index += 1
            ws[excel_alphabet[6] + str(index)] = dropdownText[i]  
            ws[excel_alphabet[5]+str(index)] = dropdownNumber[i]
            ws[excel_alphabet[3]+str(index)] = dropdownLabel[i]
            ask_turn = False 
            if (i == 0):
                ws[excel_alphabet[2]+str(index)] = "ModA_" + new_strip

    #tip
    if paragpraph.text.lstrip().find(tip_string.lstrip()) == 0:
        tip_turn = True
        index += 1
        ws[excel_alphabet[2]+str(index)] = "ModA_" + new_strip + "_Label"
        text = "<table> \n <tr> \n <td valign = top width = 700> <b> Tip: </b> \n <p><i>" + paragpraph.text.lstrip() + "</i></p><br> \n <p><b> Comment </b></p> \n </td> \n </tr> \n </table>"
        ws[excel_alphabet[6] + str(index)] = '\n'.join(['\n'.join(wrap(block, width=50)) for block in text.splitlines()])
        ws[excel_alphabet[3]+str(index)] = "Label"
        ws[excel_alphabet[1]+str(index)] = new_strip + "_Label_2"
        lable_two = True
    #memo
    if tip_turn == True:
        index += 1
        ws[excel_alphabet[2]+str(index)] = "ModA_" + new_strip + "_DESC"        
        label_count += 1
        ws[excel_alphabet[3]+str(index)] = "Memo"
        ws[excel_alphabet[4]+str(index)] = "500"
        ws[excel_alphabet[3]+str(index)] = "Memo"
        tip_turn = False
    
    if (lable_one == True and lable_two == True):
        lable_one = False
        lable_two = False 
        
        
wb.save('bc.xlsx')


